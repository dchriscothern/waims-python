from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path

import fitz
from openpyxl import Workbook
from rapidocr_onnxruntime import RapidOCR


ROOT = Path(__file__).resolve().parents[1]
DOCS_DIR = ROOT / "docs"
OUTPUT_PATH = DOCS_DIR / "arkansas_games_manual_coding.xlsx"
PDF_GLOB = "Baha-Mar-Summer-Stats-*.pdf"


OPPONENT_ALIASES = {
    "bahamas": "Bahamas",
    "carleton": "Carleton",
    "columbia": "Columbia",
    "calgary": "Calgary",
    "toros del valle": "Toros Del Valle",
    "toros-del-valle": "Toros Del Valle",
    "toros de valle": "Toros Del Valle",
}


def normalize_opponent(name: str) -> str:
    value = (name or "").strip()
    if not value:
        return "Unknown opponent"
    if value.lower().startswith("vs-"):
        value = value[3:]
    value = value.replace("_", " ")
    value = value.replace("-", " ")
    value = re.sub(r"\s+", " ", value).strip()
    normalized = value.title()
    lookup = normalized.lower()
    return OPPONENT_ALIASES.get(lookup, normalized)


def _extract_ocr_lines(pdf_path: Path) -> list[str]:
    ocr = RapidOCR()
    doc = fitz.open(str(pdf_path))
    lines: list[str] = []

    for page_index in range(doc.page_count):
        page = doc[page_index]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        temp_image = ROOT / f"__arkansas_tmp_page_{page_index}.png"
        pix.save(str(temp_image))
        try:
            result, _ = ocr(str(temp_image))
        finally:
            if temp_image.exists():
                temp_image.unlink(missing_ok=True)

        if not result:
            continue

        for item in result:
            if isinstance(item, (list, tuple)) and len(item) > 1:
                text = str(item[1]).strip()
                if text:
                    lines.append(text)

    doc.close()
    return lines


def _extract_opponent_from_text(text_blob: str, pdf_path: Path) -> str:
    for token in pdf_path.stem.split("-"):
        if token.lower() == "vs":
            continue
        if token.lower() in {"summer", "stats", "baha", "mar"}:
            continue

    if "vs" in pdf_path.stem.lower():
        candidate = pdf_path.stem.split("vs", 1)[1]
        return normalize_opponent(candidate)

    for pattern in [
        r"(Bahamas|Carleton|Columbia|Calgary|Toros\s+del\s+Valle|Toros\s+Del\s+Valle)",
        r"(Bahamas|Carleton|Columbia|Calgary|Toros.*Valle)",
    ]:
        match = re.search(pattern, text_blob, flags=re.IGNORECASE)
        if match:
            return normalize_opponent(match.group(1))
    return "Unknown opponent"


def _extract_final_score(text_blob: str, opponent: str) -> str:
    lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
    opponent_key = opponent.lower()
    candidate_lines = []
    for line in lines:
        lowered = line.lower()
        if "arkansas" in lowered or "ark" in lowered:
            candidate_lines.append(line)
        elif opponent_key in lowered or "bahamas" in lowered or "carleton" in lowered or "columbia" in lowered or "calgary" in lowered or "toros" in lowered:
            candidate_lines.append(line)

    # Prefer explicit score lines like "Arkansas-106" and "Bahamas-59" or "ARK 105-48".
    for line in candidate_lines:
        cleaned = re.sub(r"\s+", " ", line)
        m = re.search(r"(?i)(?:arkansas|ark)[^\d]*(\d{2,3})\s*[-: ]\s*(\d{2,3})", cleaned)
        if m:
            return f"{int(m.group(1))}-{int(m.group(2))}"
        m = re.search(r"(?i)(?:bahamas|carleton|columbia|calgary|toros[^\n]*valle)[^\d]*(\d{2,3})\s*[-: ]\s*(\d{2,3})", cleaned)
        if m:
            return f"{int(m.group(2))}-{int(m.group(1))}"

    for line in candidate_lines:
        cleaned = re.sub(r"\s+", " ", line)
        ark_match = re.search(r"(?i)(?:arkansas|ark)[^0-9]*(\d{2,3})", cleaned)
        opp_match = re.search(rf"(?i)(?:{re.escape(opponent)})[^0-9]*(\d{{2,3}})", cleaned)
        if ark_match and opp_match:
            return f"{int(ark_match.group(1))}-{int(opp_match.group(1))}"

    # Fallback: search for Arkansas score + opponent score anywhere in the OCR blob.
    ark_match = re.search(r"(?i)(?:arkansas|ark)[^0-9]*(\d{2,3})", text_blob)
    opp_match = re.search(rf"(?i)(?:{re.escape(opponent)})[^0-9]*(\d{{2,3}})", text_blob)
    if ark_match and opp_match:
        return f"{int(ark_match.group(1))}-{int(opp_match.group(1))}"

    return ""


def parse_game_metadata(pdf_path: Path, raw_lines: list[str]) -> dict:
    text_blob = "\n".join(raw_lines)
    date_match = re.search(r"(\d{2}/\d{2}/\d{2})", text_blob)
    date_value = date_match.group(1) if date_match else ""
    opponent = _extract_opponent_from_text(text_blob, pdf_path)
    final_score = _extract_final_score(text_blob, opponent)

    return {
        "game_id": pdf_path.stem,
        "date": date_value,
        "opponent": opponent,
        "team": "Arkansas",
        "source_pdf": str(pdf_path.name),
        "status": "OCR verified",
        "notes": "Scanned PDF converted with OCR and parsed into summary rows.",
        "final_score": final_score,
    }


def parse_game_summary_rows(docs_dir: Path = DOCS_DIR) -> list[dict]:
    rows: list[dict] = []
    for pdf_path in sorted(docs_dir.glob(PDF_GLOB)):
        lines = _extract_ocr_lines(pdf_path)
        metadata = parse_game_metadata(pdf_path, lines)
        if not metadata["game_id"]:
            continue
        rows.append(metadata)
    return rows


def import_game_summary_rows_to_db(rows: list[dict], db_path: Path | None = None) -> int:
    target = db_path or ROOT / "waims-mens" / "data" / "waims_arkansas.db"
    conn = sqlite3.connect(str(target))
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS game_results (
            game_id TEXT PRIMARY KEY,
            date TEXT,
            opponent TEXT,
            team TEXT,
            arkansas_score INTEGER,
            opponent_score INTEGER,
            final_score TEXT,
            result TEXT,
            source_pdf TEXT,
            status TEXT,
            notes TEXT
        )
        """
    )

    imported = 0
    for row in rows:
        final_score = (row.get("final_score") or "").strip()
        if not final_score or "-" not in final_score:
            continue
        arkansas_score_text, opponent_score_text = final_score.split("-", 1)
        try:
            arkansas_score = int(arkansas_score_text)
            opponent_score = int(opponent_score_text)
        except ValueError:
            continue
        result = "W" if arkansas_score > opponent_score else "L" if arkansas_score < opponent_score else "T"
        cursor.execute(
            """
            INSERT OR REPLACE INTO game_results (
                game_id, date, opponent, team, arkansas_score, opponent_score,
                final_score, result, source_pdf, status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["game_id"],
                row.get("date", ""),
                row.get("opponent", "Unknown opponent"),
                row.get("team", "Arkansas"),
                arkansas_score,
                opponent_score,
                final_score,
                result,
                row.get("source_pdf", ""),
                row.get("status", "OCR verified"),
                row.get("notes", ""),
            ),
        )
        imported += 1

    conn.commit()
    conn.close()
    return imported


def build_workbook(pdf_paths: list[Path]) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Game Summary"
    summary_ws.append(["game_id", "date", "opponent", "team", "source_pdf", "status", "final_score", "notes"])

    raw_ws = wb.create_sheet("Raw OCR")
    raw_ws.append(["game_id", "page", "line_number", "text"])

    coding_ws = wb.create_sheet("Manual Coding Template")
    coding_ws.append([
        "game_id",
        "date",
        "opponent",
        "period",
        "clock",
        "team",
        "player_name",
        "player_number",
        "event_type",
        "action",
        "points",
        "assist",
        "turnover",
        "rebound",
        "foul",
        "notes",
        "needs_review",
    ])

    for pdf_path in sorted(pdf_paths):
        raw_lines = _extract_ocr_lines(pdf_path)
        metadata = parse_game_metadata(pdf_path, raw_lines)
        summary_ws.append([
            metadata["game_id"],
            metadata["date"],
            metadata["opponent"],
            metadata["team"],
            metadata["source_pdf"],
            metadata["status"],
            metadata["final_score"],
            metadata["notes"],
        ])

        doc = fitz.open(str(pdf_path))
        for page_index in range(doc.page_count):
            page = doc[page_index]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            temp_image = ROOT / f"__arkansas_tmp_page_{page_index}.png"
            pix.save(str(temp_image))
            try:
                result, _ = RapidOCR()(str(temp_image))
            finally:
                if temp_image.exists():
                    temp_image.unlink(missing_ok=True)

            if not result:
                continue

            for line_number, item in enumerate(result, start=1):
                if isinstance(item, (list, tuple)) and len(item) > 1:
                    text = str(item[1]).strip()
                    if text:
                        raw_ws.append([metadata["game_id"], page_index + 1, line_number, text])
                        coding_ws.append([
                            metadata["game_id"],
                            metadata["date"],
                            metadata["opponent"],
                            "",
                            "",
                            metadata["team"],
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            text,
                            "Y",
                        ])
        doc.close()

    for ws in (summary_ws, raw_ws, coding_ws):
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_PATH)
    print(f"Workbook saved: {OUTPUT_PATH}")
    print(f"Games processed: {len(pdf_paths)}")


def main() -> None:
    pdf_files = sorted(DOCS_DIR.glob(PDF_GLOB))
    if not pdf_files:
        raise FileNotFoundError(f"No files matching '{PDF_GLOB}' found in {DOCS_DIR}")

    rows = parse_game_summary_rows(DOCS_DIR)
    imported = import_game_summary_rows_to_db(rows, ROOT / "waims-mens" / "data" / "waims_arkansas.db")
    print(f"Imported Arkansas game summaries: {imported}")
    build_workbook(pdf_files)


if __name__ == "__main__":
    main()
